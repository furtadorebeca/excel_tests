import openpyxl 

wb = openpyxl.load_workbook('videogamesales.xlsx')

ws = wb.active

# ws = wb['vgsales'] (this is another way to select a specific sheet)

# 
# print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))


#print('The value in cell A1 is: '+ws['A1'].value)


#values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
#print(values)

# write data to a specific cell
# data=[ws.cell(row=i,column=2).value for i in range(2,12)]
# print(data)

# reading data from a range of cells (from column 1 to 6)

# my_list = list()

# for value in ws.iter_rows(
#     min_row=1, max_row=11, min_col=1, max_col=6, 
#     values_only=True):
#     my_list.append(value)
    
# for ele1,ele2,ele3,ele4,ele5,ele6 in my_list:
#     (print ("{:<8}{:<35}{:<10}{:<10}{:<15}{:<15}".format(ele1,ele2,ele3,ele4,ele5,ele6)))

# ws['K1'] = 'Sum of Sales'

# ws.cell(row=1, column=11, value = 'Sum of Sales')
# wb.save('videogamesales.xlsx')

# row_position = 2
# col_position = 7

# total_sales = ((ws.cell(row=row_position, column=col_position).value)+
#                (ws.cell(row=row_position, column=col_position+1).value)+
#                (ws.cell(row=row_position, column=col_position+2).value)+
#                (ws.cell(row=row_position, column=col_position+3).value))

# ws.cell(row=2,column=11).value=total_sales
# wb.save('videogamesales.xlsx')

# Calculate total sales for each row and write to column 11
# Start from row 2 to avoid headers
# and assuming sales data is in columns 7 to 10 (NA_Sales to Other_Sales)
# Initialize row position

row_position = 1

for i in range(1, ws.max_row):

    row_position += 1
    NA_Sales = ws.cell(row=row_position, column=7).value
    EU_Sales = ws.cell(row=row_position, column=8).value
    JP_Sales = ws.cell(row=row_position, column=9).value
    Other_Sales = ws.cell(row=row_position, column=10).value

    total_sales = (NA_Sales + EU_Sales + JP_Sales + Other_Sales)
    ws.cell(row=row_position, column=11).value = total_sales

wb.save("videogamesales.xlsx")